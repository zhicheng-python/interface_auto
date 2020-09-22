# encoding:utf-8
# author:wangzhicheng
# time:2019/11/22 14:47
# file:handle_excel.py

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, colors
from Test_scripts.handle_path import do_path
from Test_scripts.handle_read_config import do_read_yaml


class DataObj:
    pass


class ReadExcel:
    """读取excel数据，写回测试数据"""

    def __init__(self, data_path=do_path.data_path, sheet_name=do_read_yaml.read_config("excel", "sheet1_name")):

        self.path = data_path
        self.sheet_name = sheet_name

    def open_data(self):  # 打开文件，创建文件对象，表单对象
        self.wb = load_workbook(self.path)
        self.sheet = self.wb[self.sheet_name]

    def get_title(self):  # 获取数据表头
        self.open_data()  # 调用实例方法打开文件

        # 列表推导式sheet.rows 按行 获取所有数据，每行数据为一个元素存放在列表中，列表[0],取第一行表头的值.value
        title = [i.value for i in list(self.sheet.rows)[0]]

        self.wb.close()  # 关闭文件
        return title

    def get_data(self):

        title = self.get_title()  # 先获取表头（此方法中有关闭文件的操作）

        self.open_data()  # 打开文件

        # sheet.rows[1:]获取 除第一行以为的所有行数据，把每一行单元格的值与表头的zip

        datas = [dict(zip(title, [j.value for j in i])) for i in list(self.sheet.rows)[1:]]  # 列表推导式

        self.wb.close()

        return datas

    def get_data_obj(self):  # 获取已设置实例属性的实例对象

        title = self.get_title()  # 先获取表头（此方法中有关闭文件的操作）

        self.open_data()  # 打开文件

        datas = []

        for i in list(self.sheet.rows)[1:]:

            data = [j.value for j in i]

            data_obj = DataObj()  # 创建一个实例对象

            for k in zip(title, data):  # zip之后的数据格式类似 [(k[0],k[1]),(k[0],k[1])...]
                setattr(data_obj, k[0], k[1])  # setattr 设置实例对象属性
            datas.append(data_obj)  # 将设置了实例属性的实例对象添加到列表中

        self.wb.close()
        return datas

    def write_data(self, row, column, value):  # 写回测试结果

        self.open_data()

        self.sheet.cell(row, column, value)  # sheet.cell 写入指定单元格数据

        self.wb.save(self.path)

        self.wb.close()

    def write_color(self):

        title = self.get_title()
        self.open_data()

        # fill = PatternFill("lightVertical", fgColor=colors.RED)
        # fill1 = PatternFill("lightVertical", fgColor=colors.GREEN)

        fill = PatternFill("solid", fgColor=colors.RED)
        fill1 = PatternFill("solid", fgColor=colors.GREEN)

        for i in list(self.sheet.columns)[title.index("result")]:
            if i.value == do_read_yaml.read_config("excel", "pass_value"):
                i.fill = fill1
            if i.value == do_read_yaml.read_config("excel", "fail_value"):
                i.fill = fill
        self.wb.save(self.path)
        self.wb.close()


do_excel = ReadExcel()

if __name__ == '__main__':
    title = do_excel.get_title()
    data = do_excel.get_data()
    datas = do_excel.get_data_obj()


    for i in ReadExcel(sheet_name="invest").get_data():
        print(i)